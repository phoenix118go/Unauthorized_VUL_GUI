import sys
import json
import signal
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
                             QPushButton, QTextEdit, QLineEdit, QFileDialog, QLabel,
                             QProgressBar, QCheckBox, QGroupBox, QFormLayout, QSpinBox,
                             QMessageBox, QSplitter, QApplication, QScrollArea)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QUrl
from PyQt6.QtGui import QColor, QTextCursor, QIcon
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import requests
import socket
import urllib3
from urllib.parse import urlparse

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


def normalize_target_input(raw_target: str) :
    """
    目标输入解析：支持IP、IP:PORT、URL、域名、域名:PORT等格式
    返回标准化信息字典（含host、port、scheme、full_url）
    """
    if not raw_target :
        return None

    # 处理纯IP:PORT格式
    if ":" in raw_target and not raw_target.startswith(("http://", "https://", "ftp://")) :
        parts = raw_target.split(":", 1)  # 只分割第一个冒号（避免IPv6冲突）
        if len(parts) == 2 and parts[1].isdigit() :
            host, port = parts[0], int(parts[1])
            return {
                "host" : host,
                "port" : port,
                "scheme" : "http",
                "full_url" : f"http://{host}:{port}"
            }

    # 处理URL或域名格式
    if not raw_target.startswith(("http://", "https://")) :
        raw_target = "http://" + raw_target  # 补全协议

    parsed = urlparse(raw_target)
    return {
        "host" : parsed.hostname or raw_target,  # 兼容纯IP输入
        "port" : parsed.port,
        "scheme" : parsed.scheme or "http",
        "full_url" : raw_target
    }


# ---------------------- 2. 漏洞扫描核心类（含速率优化） ----------------------
class VulnerabilityScanner :
    def __init__(self, proxy=None, timeout=5) :  # 缩短超时时间（可配置）
        self.proxy = proxy
        self.timeout = timeout
        # 复用会话（连接池优化）
        self.session = requests.Session()
        self.session.mount('http://', requests.adapters.HTTPAdapter(pool_connections=20, pool_maxsize=20))
        self.session.mount('https://', requests.adapters.HTTPAdapter(pool_connections=20, pool_maxsize=20))

        if proxy :
            self.session.proxies = {"http" : proxy, "https" : proxy, "socks": proxy,"socks5": proxy}
        self.session.verify = False
        self.session.headers = {
            "User-Agent" : "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36"
        }

        # 服务检测器和默认端口（保持原有映射）
        self.detectors = {
            "ftp" : self.check_ftp,
            "redis" : self.check_redis,
            "docker" : self.check_docker,
            "docker_registry" : self.check_docker_registry,
            "elasticsearch" : self.check_elasticsearch,
            "jenkins" : self.check_jenkins,
            "kibana" : self.check_kibana,
            "zookeeper" : self.check_zookeeper,
            "mongodb" : self.check_mongodb,
            "kubernetes" : self.check_kubernetes,
            "jupyter" : self.check_jupyter,
            "nacos" : self.check_nacos,
            "ollama" : self.check_ollama,
            "rsync" : self.check_rsync,
            "swagger" : self.check_swagger,
            "springboot" : self.check_springboot,
            "druid" : self.check_druid,
            "ldap" : self.check_ldap,
            "vnc" : self.check_vnc,
            "couchdb" : self.check_couchdb,
            "spark" : self.check_spark,
            "weblogic" : self.check_weblogic,
            "hadoop" : self.check_hadoop,
            "jboss" : self.check_jboss,
            "activemq" : self.check_activemq,
            "zabbix" : self.check_zabbix,
            "memcached" : self.check_memcached,
            "rabbitmq" : self.check_rabbitmq,
            "nfs" : self.check_nfs,
            "dubbo" : self.check_dubbo,
            "solr" : self.check_solr,
            "harbor" : self.check_harbor,
            "smb" : self.check_smb,
            "wordpress" : self.check_wordpress,
            "crowd" : self.check_crowd,
            "uwsgi" : self.check_uwsgi,
            "kong" : self.check_kong,
            "thinkadmin" : self.check_thinkadmin
        }

        self.default_ports = {
            "ftp" : 21,
            "redis" : 6379,
            "docker" : 2375,
            "docker_registry" : 5000,
            "elasticsearch" : 9200,
            "jenkins" : 8080,
            "kibana" : 5601,
            "zookeeper" : 2181,
            "mongodb" : 27017,
            "kubernetes" : 8080,
            "jupyter" : 8888,
            "nacos" : 8848,
            "ollama" : 11434,
            "rsync" : 873,
            "swagger" : 80,
            "springboot" : 8080,
            "druid" : 8080,
            "ldap" : 389,
            "vnc" : 5900,
            "couchdb" : 5984,
            "spark" : 6066,
            "weblogic" : 7001,
            "hadoop" : 8088,
            "jboss" : 8080,
            "activemq" : 8161,
            "zabbix" : 10051,
            "memcached" : 11211,
            "rabbitmq" : 15672,
            "nfs" : 2049,
            "dubbo" : 28096,
            "solr" : 8983,
            "harbor" : 80,
            "smb" : 445,
            "wordpress" : 80,
            "crowd" : 8095,
            "uwsgi" : 1717,
            "kong" : 8001,
            "thinkadmin" : 80
        }

        # 非HTTP协议服务列表
        self.NON_HTTP_SERVICES = {
            "smb", "ftp", "redis", "zookeeper", "mongodb", "ldap",
            "vnc", "memcached", "nfs", "dubbo", "rsync", "uwsgi"
        }

    # ---------------------- 检测方法（含速率优化） ----------------------
    def check_ftp(self, target_info) :
        host = target_info["host"]
        port = target_info.get("port", 21)
        try :
            with socket.create_connection((host, port), timeout=self.timeout) as sock :
                sock.sendall(b"USER anonymous\r\n")
                response = sock.recv(1024).decode(errors="ignore")
                if "331" in response :
                    sock.sendall(b"PASS anonymous@example.com\r\n")
                    response = sock.recv(1024).decode(errors="ignore")
                    if "230" in response :
                        return True, "FTP anonymous login successful"
        except Exception as e :
            return False, f"检测失败: {str(e)}"
        return False, "未发现未授权访问"

    def check_redis(self, target_info) :
        host = target_info["host"]
        port = target_info.get("port", 6379)
        try :
            with socket.create_connection((host, port), timeout=self.timeout) as sock :
                sock.sendall(b"INFO\r\n")
                response = sock.recv(1024).decode(errors="ignore")
                if "redis_version" in response :
                    return True, "Redis未授权访问"
        except Exception as e :
            return False, f"检测失败: {str(e)}"
        return False, "未发现未授权访问"

    def check_docker(self, target_info) :
        """检测Docker未授权访问"""
        host = target_info["host"]
        port = target_info.get("port", 2375)
        scheme = target_info.get("scheme", "http")
        url = f"{scheme}://{host}:{port}/version"
        try :
            response = self.session.get(url, timeout=self.timeout)
            if response.status_code == 200 and "ApiVersion" in response.text :
                return True, f"Docker unauthorized access"
        except Exception as e :
            return False, f"Docker detection failed: {str(e)}"
        return False, "Docker unauthorized access not found"

    def check_docker_registry(self, target_info):
        """检测Docker Registry未授权访问"""
        host = target_info["host"]
        port = target_info.get("port", 5000)
        scheme = target_info.get("scheme", "http")
        url = f"{scheme}://{host}:{port}/v2/_catalog"
        try:
            response = self.session.get(url, timeout=self.timeout)
            if response.status_code == 200 and "repositories" in response.text:
                return True, "Docker Registry unauthorized access"
        except Exception as e:
            return False, f"Docker Registry detection failed: {str(e)}"
        return False, "Docker Registry unauthorized access not found"

    def check_elasticsearch(self, target_info):
        """检测Elasticsearch未授权访问"""
        host = target_info["host"]
        port = target_info.get("port", 9200)
        scheme = target_info.get("scheme", "http")
        url = f"{scheme}://{host}:{port}/_cat/indices"
        try:
            response = self.session.get(url, timeout=self.timeout)
            if response.status_code == 200 or "green" in response.text or "yellow" in response.text:
                return True, "Elasticsearch unauthorized access"
        except Exception as e:
            return False, f"Elasticsearch detection failed: {str(e)}"
        return False, "Elasticsearch unauthorized access not found"

    def check_jenkins(self, target_info):
        """检测Jenkins未授权访问"""
        host = target_info["host"]
        port = target_info.get("port", 8080)
        scheme = target_info.get("scheme", "http")
        url = f"{scheme}://{host}:{port}/api/json"
        try:
            response = self.session.get(url, timeout=self.timeout)
            if response.status_code == 200 and "jobs" in response.text:
                return True, "Jenkins unauthorized access"
        except Exception as e:
            return False, f"Jenkins detection failed: {str(e)}"
        return False, "Jenkins unauthorized access not found"

    def check_kibana(self, target_info):
        """检测Kibana未授权访问"""
        host = target_info["host"]
        port = target_info.get("port", 5601)
        scheme = target_info.get("scheme", "http")
        url = f"{scheme}://{host}:{port}/api/status"
        try:
            response = self.session.get(url, timeout=self.timeout)
            if response.status_code == 200 and "status" in response.text:
                return True, "Kibana unauthorized access"
        except Exception as e:
            return False, f"Kibana detection failed: {str(e)}"
        return False, "Kibana unauthorized access not found"

    def check_zookeeper(self, target_info):
        """检测Zookeeper未授权访问"""
        host = target_info["host"]
        port = target_info.get("port", 2181)
        try:
            with socket.create_connection((host, port), timeout=self.timeout) as sock:
                sock.sendall(b"stat\r\n")
                response = sock.recv(1024).decode(errors="ignore")
                if "Zookeeper version" in response:
                    return True, "Zookeeper unauthorized access"
        except Exception as e:
            return False, f"Zookeeper detection failed: {str(e)}"
        return False, "Zookeeper unauthorized access not found"

    def check_mongodb(self, target_info):
        """检测MongoDB未授权访问"""
        host = target_info["host"]
        port = target_info.get("port", 27017)
        try:
            with socket.create_connection((host, port), timeout=self.timeout) as sock:
                sock.sendall(b"db.adminCommand('ping')\r\n")
                response = sock.recv(1024).decode(errors="ignore")
                if "ok" in response:
                    return True, "MongoDB unauthorized access"
        except Exception as e:
            return False, f"MongoDB detection failed: {str(e)}"
        return False, "MongoDB unauthorized access not found"

    def check_kubernetes(self, target_info):
        """检测Kubernetes未授权访问"""
        host = target_info["host"]
        port = target_info.get("port", 8080)
        scheme = target_info.get("scheme", "http")
        url = f"{scheme}://{host}:{port}/api/v1/namespaces/default/pods"
        try:
            response = self.session.get(url, timeout=self.timeout)
            if response.status_code == 200 and "items" in response.text:
                return True, "Kubernetes unauthorized access"
        except Exception as e:
            return False, f"Kubernetes detection failed: {str(e)}"
        return False, "Kubernetes unauthorized access not found"

    def check_jupyter(self, target_info):
        """检测Jupyter Notebook未授权访问"""
        host = target_info["host"]
        port = target_info.get("port", 8888)
        scheme = target_info.get("scheme", "http")
        url = f"{scheme}://{host}:{port}/api/kernels"
        try:
            response = self.session.get(url, timeout=self.timeout)
            if response.status_code == 200 and "kernels" in response.text:
                return True, "Jupyter Notebook unauthorized access"
        except Exception as e:
            return False, f"Jupyter detection failed: {str(e)}"
        return False, "Jupyter unauthorized access not found"

    def check_nacos(self, target_info):
        """检测Nacos未授权访问"""
        host = target_info["host"]
        port = target_info.get("port", 8848)
        scheme = target_info.get("scheme", "http")
        url = f"{scheme}://{host}:{port}/nacos/v1/auth/users?pageNo=1&pageSize=10"
        try:
            response = self.session.get(url, timeout=self.timeout)
            if response.status_code == 200 and "username" in response.text:
                return True, "Nacos unauthorized access"
        except Exception as e:
            return False, f"Nacos detection failed: {str(e)}"
        return False, "Nacos unauthorized access not found"

    def check_ollama(self, target_info):
        """检测Ollama未授权访问"""
        host = target_info["host"]
        port = target_info.get("port", 11434)
        scheme = target_info.get("scheme", "http")
        url = f"{scheme}://{host}:{port}/api/tags"
        try:
            response = self.session.get(url, timeout=self.timeout)
            if response.status_code == 200 and "models" in response.text:
                return True, "Ollama unauthorized access"
        except Exception as e:
            return False, f"Ollama detection failed: {str(e)}"
        return False, "Ollama unauthorized access not found"

    def check_rsync(self, target_info):
        """检测Rsync未授权访问"""
        host = target_info["host"]
        port = target_info.get("port", 873)
        try:
            with socket.create_connection((host, port), timeout=self.timeout) as sock:
                sock.sendall(b"@RSYNCD: 31.0\n")
                response = sock.recv(1024).decode(errors="ignore")
                if "RSYNCD" in response:
                    return True, "Rsync unauthorized access"
        except Exception as e:
            return False, f"Rsync detection failed: {str(e)}"
        return False, "Rsync unauthorized access not found"

    def check_swagger(self, target_info):
        """Swagger UI未授权访问检测"""
        host = target_info["host"]
        port = target_info.get("port", 80)
        scheme = target_info.get("scheme", "http")

        # 常见的Swagger UI路径
        paths = [
            "/swagger-ui.html",
            "/swagger/index.html",
            "/swagger/ui/index",
            "/swagger",
            "/api-docs",
            "/v2/api-docs",
            "/swagger-resources",
            "/swagger-ui",
            "/api/swagger-ui.html",
            "/docs",
            "/swagger-ui/index.html"
        ]

        # 尝试所有可能的路径
        for path in paths:
            url = f"{scheme}://{host}:{port}{path}" if port else f"{scheme}://{host}{path}"
            try:
                response = self.session.get(url, timeout=self.timeout)
                if response.status_code == 200:
                    # 使用更精确的识别方法
                    if any(keyword in response.text for keyword in
                           ["Swagger UI", "swagger-ui", "swagger.json", "swagger.yaml"]):
                        return True, f"Swagger UI unauthorized access (path: {path})"
            except Exception:
                continue

        return False, "Swagger unauthorized access not found"

    def check_springboot(self, target_info):
        """检测SpringBoot Actuator未授权访问"""
        host = target_info["host"]
        port = target_info.get("port", 80)
        scheme = target_info.get("scheme", "http")

        # 常见的Actuator路径
        paths = [
            "/actuator",
            "/actuator/health",
            "/actuator/env",
            "/actuator/metrics",
            "/actuator/beans",
            "/actuator/mappings"
        ]

        for path in paths:
            url = f"{scheme}://{host}:{port}{path}" if port else f"{scheme}://{host}{path}"
            try:
                response = self.session.get(url, timeout=self.timeout)
                if response.status_code == 200 and "actuator" in response.text:
                    return True, f"SpringBoot Actuator unauthorized access (path: {path})"
            except Exception:
                continue

        return False, "SpringBoot unauthorized access not found"

    def check_druid(self, target_info):
        """检测Druid未授权访问"""
        host = target_info["host"]
        port = target_info.get("port", 80)
        scheme = target_info.get("scheme", "http")

        # 常见的Druid路径
        paths = [
            "/druid/index.html",
            "/druid/login.html",
            "/druid/weburi.html",
            "/druid/websession.html",
            "/druid/sql.html"
        ]

        for path in paths:
            url = f"{scheme}://{host}:{port}{path}" if port else f"{scheme}://{host}{path}"
            try:
                response = self.session.get(url, timeout=self.timeout)
                if response.status_code == 200 and "Druid" in response.text:
                    return True, f"Druid unauthorized access (path: {path})"
            except Exception:
                continue

        return False, "Druid unauthorized access not found"

    def check_ldap(self, target_info):
        """检测LDAP未授权访问"""
        host = target_info["host"]
        port = target_info.get("port", 389)
        try:
            with socket.create_connection((host, port), timeout=self.timeout) as sock:
                # LDAP匿名绑定尝试
                bind_request = bytes.fromhex("30 0c 02 01 01 60 07 02 01 03 04 00 80 00")
                sock.sendall(bind_request)
                response = sock.recv(1024)
                if response and len(response) > 0:
                    return True, "LDAP unauthorized access (anonymous bind possible)"
        except Exception as e:
            return False, f"LDAP detection failed: {str(e)}"
        return False, "LDAP unauthorized access not found"

    def check_vnc(self, target_info):
        """检测VNC未授权访问"""
        host = target_info["host"]
        port = target_info.get("port", 5900)
        try:
            with socket.create_connection((host, port), timeout=self.timeout) as sock:
                # 读取VNC协议响应
                response = sock.recv(1024)
                if b"RFB" in response:
                    return True, "VNC service exposed (possible unauthorized access)"
        except Exception as e:
            return False, f"VNC detection failed: {str(e)}"
        return False, "VNC unauthorized access not found"

    def check_couchdb(self, target_info):
        """检测CouchDB未授权访问"""
        host = target_info["host"]
        port = target_info.get("port", 5984)
        scheme = target_info.get("scheme", "http")
        url = f"{scheme}://{host}:{port}/_all_dbs"
        try:
            response = self.session.get(url, timeout=self.timeout)
            if response.status_code == 200 and "[" in response.text:
                return True, "CouchDB unauthorized access"
        except Exception as e:
            return False, f"CouchDB detection failed: {str(e)}"
        return False, "CouchDB unauthorized access not found"

    def check_spark(self, target_info):
        """检测Apache Spark未授权访问"""
        host = target_info["host"]
        port = target_info.get("port", 6066)
        scheme = target_info.get("scheme", "http")
        url = f"{scheme}://{host}:{port}/"
        try:
            response = self.session.get(url, timeout=self.timeout)
            if response.status_code == 200 and "Spark" in response.text:
                return True, "Apache Spark unauthorized access"
        except Exception as e:
            return False, f"Apache Spark detection failed: {str(e)}"
        return False, "Apache Spark unauthorized access not found"

    def check_weblogic(self, target_info):
        """检测Weblogic未授权访问"""
        host = target_info["host"]
        port = target_info.get("port", 7001)
        scheme = target_info.get("scheme", "http")
        url = f"{scheme}://{host}:{port}/console/login/LoginForm.jsp"
        try:
            response = self.session.get(url, timeout=self.timeout)
            if response.status_code == 200 and "WebLogic Server" in response.text:
                return True, "Weblogic console exposed (possible unauthorized access)"
        except Exception as e:
            return False, f"Weblogic detection failed: {str(e)}"
        return False, "Weblogic unauthorized access not found"

    def check_hadoop(self, target_info):
        """检测Hadoop YARN未授权访问"""
        host = target_info["host"]
        port = target_info.get("port", 8088)
        scheme = target_info.get("scheme", "http")
        url = f"{scheme}://{host}:{port}/ws/v1/cluster/apps"
        try:
            response = self.session.get(url, timeout=self.timeout)
            if response.status_code == 200 and "apps" in response.text:
                return True, "Hadoop YARN unauthorized access"
        except Exception as e:
            return False, f"Hadoop detection failed: {str(e)}"
        return False, "Hadoop unauthorized access not found"

    def check_jboss(self, target_info):
        """检测JBoss未授权访问"""
        host = target_info["host"]
        port = target_info.get("port", 8080)
        scheme = target_info.get("scheme", "http")
        url = f"{scheme}://{host}:{port}/jmx-console/"
        try:
            response = self.session.get(url, timeout=self.timeout)
            if response.status_code == 200 and "JBoss" in response.text:
                return True, "JBoss JMX console exposed (possible unauthorized access)"
        except Exception as e:
            return False, f"JBoss detection failed: {str(e)}"
        return False, "JBoss unauthorized access not found"

    def check_activemq(self, target_info):
        """检测ActiveMQ未授权访问"""
        host = target_info["host"]
        port = target_info.get("port", 8161)
        scheme = target_info.get("scheme", "http")
        url = f"{scheme}://{host}:{port}/admin/"
        try:
            response = self.session.get(url, timeout=self.timeout)
            if response.status_code == 200 and "ActiveMQ" in response.text:
                return True, "ActiveMQ management console exposed (possible unauthorized access)"
        except Exception as e:
            return False, f"ActiveMQ detection failed: {str(e)}"
        return False, "ActiveMQ unauthorized access not found"

    def check_zabbix(self, target_info):
        """检测Zabbix未授权访问"""
        host = target_info["host"]
        port = target_info.get("port", 10051)
        scheme = target_info.get("scheme", "http")
        url = f"{scheme}://{host}:{port}/"
        try:
            response = self.session.get(url, timeout=self.timeout)
            if response.status_code == 200 and "Zabbix" in response.text:
                return True, "Zabbix service exposed (possible unauthorized access)"
        except Exception as e:
            return False, f"Zabbix detection failed: {str(e)}"
        return False, "Zabbix unauthorized access not found"

    def check_memcached(self, target_info):
        """检测Memcached未授权访问"""
        host = target_info["host"]
        port = target_info.get("port", 11211)
        try:
            with socket.create_connection((host, port), timeout=self.timeout) as sock:
                sock.sendall(b"stats\r\n")
                response = sock.recv(1024).decode(errors="ignore")
                if "STAT" in response:
                    return True, "Memcached unauthorized access"
        except Exception as e:
            return False, f"Memcached detection failed: {str(e)}"
        return False, "Memcached unauthorized access not found"

    def check_rabbitmq(self, target_info):
        """检测RabbitMQ未授权访问"""
        host = target_info["host"]
        port = target_info.get("port", 15672)
        scheme = target_info.get("scheme", "http")
        url = f"{scheme}://{host}:{port}/api/overview"
        try:
            response = self.session.get(url, timeout=self.timeout)
            if response.status_code == 200 and "management_version" in response.text:
                return True, "RabbitMQ management API unauthorized access"
        except Exception as e:
            return False, f"RabbitMQ detection failed: {str(e)}"
        return False, "RabbitMQ unauthorized access not found"

    def check_nfs(self, target_info):
        """检测NFS未授权访问"""
        host = target_info["host"]
        port = target_info.get("port", 2049)
        try:
            with socket.create_connection((host, port), timeout=self.timeout) as sock:
                sock.sendall(b"\x80\x00\x00\x00")
                response = sock.recv(1024)
                if response and len(response) > 0:
                    return True, "NFS service exposed (possible unauthorized access)"
        except Exception as e:
            return False, f"NFS detection failed: {str(e)}"
        return False, "NFS unauthorized access not found"

    def check_dubbo(self, target_info):
        """检测Dubbo未授权访问"""
        host = target_info["host"]
        port = target_info.get("port", 28096)
        try:
            with socket.create_connection((host, port), timeout=self.timeout) as sock:
                sock.sendall(b"ls\r\n")
                response = sock.recv(1024).decode(errors="ignore")
                if "dubbo>" in response:
                    return True, "Dubbo console unauthorized access"
        except Exception as e:
            return False, f"Dubbo detection failed: {str(e)}"
        return False, "Dubbo unauthorized access not found"

    def check_solr(self, target_info):
        """检测Solr未授权访问"""
        host = target_info["host"]
        port = target_info.get("port", 8983)
        scheme = target_info.get("scheme", "http")
        url = f"{scheme}://{host}:{port}/solr/admin/cores"
        try:
            response = self.session.get(url, timeout=self.timeout)
            if response.status_code == 200 and "responseHeader" in response.text:
                return True, "Solr unauthorized access"
        except Exception as e:
            return False, f"Solr detection failed: {str(e)}"
        return False, "Solr unauthorized access not found"

    def check_harbor(self, target_info):
        """检测Harbor未授权添加管理员漏洞"""
        host = target_info["host"]
        port = target_info.get("port", 80)
        scheme = target_info.get("scheme", "http")
        url = f"{scheme}://{host}:{port}/api/v2.0/users"
        try:
            response = self.session.get(url, timeout=self.timeout)
            if response.status_code == 200 and "username" in response.text:
                return True, "Harbor user API exposed (possible unauthorized admin addition)"
        except Exception as e:
            return False, f"Harbor detection failed: {str(e)}"
        return False, "Harbor unauthorized access not found"

    def check_smb(self, target_info):
        """检测Windows共享未授权访问"""
        host = target_info["host"]
        port = target_info.get("port", 445)
        try:
            with socket.create_connection((host, port), timeout=self.timeout) as sock:
                # 发送SMB协商请求
                sock.sendall(b"\x00\x00\x00\x85\xff\x53\x4d\x42\x72\x00\x00\x00\x00\x18\x53\xc8")
                response = sock.recv(1024)
                if response and len(response) > 0:
                    return True, "SMB service exposed (possible unauthorized access)"
        except Exception as e:
            return False, f"SMB detection failed: {str(e)}"
        return False, "SMB unauthorized access not found"

    def check_wordpress(self, target_info):
        """检测WordPress未授权访问"""
        host = target_info["host"]
        port = target_info.get("port", 80)
        scheme = target_info.get("scheme", "http")
        url = f"{scheme}://{host}:{port}/wp-admin/"
        try:
            response = self.session.get(url, timeout=self.timeout)
            if response.status_code == 200 and "WordPress" in response.text:
                return True, "WordPress admin panel exposed (possible unauthorized access)"
        except Exception as e:
            return False, f"WordPress detection failed: {str(e)}"
        return False, "WordPress unauthorized access not found"

    def check_crowd(self, target_info):
        """检测Atlassian Crowd未授权访问"""
        host = target_info["host"]
        port = target_info.get("port", 8095)
        scheme = target_info.get("scheme", "http")
        url = f"{scheme}://{host}:{port}/crowd/admin/"
        try:
            response = self.session.get(url, timeout=self.timeout)
            if response.status_code == 200 and "Crowd" in response.text:
                return True, "Crowd management console exposed (possible unauthorized access)"
        except Exception as e:
            return False, f"Crowd detection failed: {str(e)}"
        return False, "Crowd unauthorized access not found"

    def check_uwsgi(self, target_info):
        """检测uWSGI未授权访问"""
        host = target_info["host"]
        port = target_info.get("port", 1717)
        try:
            with socket.create_connection((host, port), timeout=self.timeout) as sock:
                sock.sendall(b"add-mapping /foo /bar\n")
                response = sock.recv(1024).decode(errors="ignore")
                if "OK" in response:
                    return True, "uWSGI unauthorized access"
        except Exception as e:
            return False, f"uWSGI detection failed: {str(e)}"
        return False, "uWSGI unauthorized access not found"

    def check_kong(self, target_info):
        """检测Kong未授权访问"""
        host = target_info["host"]
        port = target_info.get("port", 8001)
        scheme = target_info.get("scheme", "http")
        url = f"{scheme}://{host}:{port}/"
        try:
            response = self.session.get(url, timeout=self.timeout)
            if response.status_code == 200 and "kong" in response.text:
                return True, "Kong management API exposed (possible unauthorized access)"
        except Exception as e:
            return False, f"Kong detection failed: {str(e)}"
        return False, "Kong unauthorized access not found"

    def check_thinkadmin(self, target_info):
        """检测ThinkAdmin未授权访问"""
        host = target_info["host"]
        port = target_info.get("port", 80)
        scheme = target_info.get("scheme", "http")
        url = f"{scheme}://{host}:{port}/admin.html"
        try:
            response = self.session.get(url, timeout=self.timeout)
            if response.status_code == 200 and "ThinkAdmin" in response.text:
                return True, "ThinkAdmin admin panel exposed (possible unauthorized access)"
        except Exception as e:
            return False, f"ThinkAdmin detection failed: {str(e)}"
        return False, "ThinkAdmin unauthorized access not found"

    def scan_single_service(self, target_info, service_name, custom_ports=None) :
        # 优先使用用户自定义端口，其次默认端口
        # 1. 获取用户输入的目标端口（URL中自带的端口）
        user_input_port = target_info.get("port")

        # 2. 获取服务自定义端口（表格中用户设置的端口）
        custom_port = custom_ports.get(service_name) if custom_ports else None

        # 3. 确定最终使用的端口：用户输入端口优先，其次是自定义端口，最后是默认端口
        if user_input_port is not None :  # 优先使用用户输入的目标端口
            port = user_input_port
        elif custom_port is not None :  # 其次使用服务自定义端口
            port = custom_port
        else :  # 最后使用服务默认端口
            port = self.default_ports.get(service_name)

        if not port :
            return False, f"服务{service_name}无可用端口（用户未输入端口且无默认配置）"

        # 使用最终确定的端口
        service_target = target_info.copy()
        service_target["port"] = port

        detector = self.detectors.get(service_name)
        if not detector :
            return False, f"不支持的服务: {service_name}"
        return detector(service_target)

    def detect_service(self, target_info, service_name):
        """检测指定服务（保留原有逻辑）"""
        detector = self.detectors.get(service_name)
        if not detector:
            return False, f"Unsupported service: {service_name}"
        return detector(target_info)

    def get_display_url(self, target_info, service_name, custom_ports=None) :
        """生成显示用的URL，包含服务实际扫描的端口"""
        # 获取实际使用的端口（自定义端口优先，其次默认端口）
        user_input_port = target_info.get("port")
        custom_port = custom_ports.get(service_name) if custom_ports else None

        if user_input_port is not None :
            port = user_input_port
        elif custom_port is not None :
            port = custom_port
        else :
            port = self.default_ports.get(service_name)

        if not port :
            return f"{target_info['host']} (端口未知)"

        # 非HTTP服务：仅返回 host:port
        if service_name in self.NON_HTTP_SERVICES :
            return f"{target_info['host']}:{port}"
        # HTTP服务：返回 scheme://host:port
        else :
            scheme = target_info["scheme"] or "http"
            return f"{scheme}://{target_info['host']}:{port}"


class ScanThread(QThread):
    # 新增：当前扫描服务日志信号
    scanning_service = pyqtSignal(str, str, str)  # 目标、服务、扫描路径
    # 新增：服务扫描结果信号
    service_result = pyqtSignal(str, str, bool, str)  # 目标、服务、是否存在漏洞、信息
    progress_updated = pyqtSignal(int)  # 精确进度（0-100）
    scan_completed = pyqtSignal(list)

    def __init__(self, targets, services, custom_ports=None, proxy=None, threads=20):
        super().__init__()
        self.targets = targets  # 标准化后的目标列表（含host、port等）
        self.services = services  # 选中的服务列表
        self.custom_ports = custom_ports or {}  # 用户自定义端口
        self.proxy = proxy
        self.threads = threads
        self.scanner = VulnerabilityScanner(proxy=proxy)
        self.running = True
        self.futures = []  # 用于跟踪线程池任务
        # 计算总任务量（目标数×服务数）
        self.total_tasks = len(self.targets) * len(self.services)
        self.completed_tasks = 0

    def run(self):
        results = []
        # 初始化每个目标的结果容器
        target_results = {t["full_url"]: {"target": t["full_url"], "vulnerabilities": []} for t in self.targets}

        with ThreadPoolExecutor(max_workers=self.threads) as executor:
            # 提交所有任务并记录futures
            for target in self.targets:
                target_url = target["full_url"]
                for service in self.services:
                    if not self.running:
                        break  # 若已停止，不再提交新任务
                    # 提交单个服务扫描任务
                    future = executor.submit(
                        self.scanner.scan_single_service,
                        target_info=target,
                        service_name=service,
                        custom_ports=self.custom_ports
                    )
                    # 绑定目标和服务信息（用于结果解析）
                    future.target = target  # 保存原始target信息
                    future.target_url = target_url
                    future.service = service
                    self.futures.append(future)

            # 处理已完成的任务
            for future in as_completed(self.futures):
                if not self.running:
                    break  # 若已停止，不再处理结果

                target_url = future.target_url
                service = future.service
                try:
                    status, message = future.result()
                    # 获取服务实际使用的端口（自定义或默认）
                    actual_port = self.custom_ports.get(service) or self.scanner.default_ports.get(service)
                    # 解析原始目标的host和scheme
                    parsed = urlparse(target_url)
                    # 重构显示用的URL（使用服务端口）
                    display_url = self.scanner.get_display_url(
                        target_info=future.target,
                        service_name=service,
                        custom_ports=self.custom_ports
                    )
                    # 发射信号时使用重构的URL
                    self.service_result.emit(display_url, service, status, message)
                    # 记录结果
                    target_results[target_url]["vulnerabilities"].append({
                        "service": service,
                        "status": status,
                        "message": message
                    })
                except Exception as e:
                    # 异常时也使用重构的URL显示
                    actual_port = self.custom_ports.get(service) or self.scanner.default_ports.get(service)
                    parsed = urlparse(target_url)
                    display_url = f"{parsed.scheme}://{parsed.hostname}:{actual_port}"
                    self.service_result.emit(display_url, service, False, f"扫描异常: {str(e)}")

                # 更新进度
                self.completed_tasks += 1
                progress = int((self.completed_tasks / self.total_tasks) * 100)
                self.progress_updated.emit(progress)

        # 收集非空结果
        final_results = [v for v in target_results.values() if v["vulnerabilities"]]
        self.scan_completed.emit(final_results)

    def stop(self):
        """立即停止所有扫描任务（强制终止未完成任务）"""
        self.running = False
        # 取消所有未完成的任务
        for future in self.futures:
            if not future.done():
                future.cancel()
        # 强制关闭线程池（不等待剩余任务）
        self.requestInterruption()


class UnauthorizedScanTab(QWidget) :
    def __init__(self) :
        super().__init__()
        self.init_ui()
        self.scan_thread = None

    def init_ui(self) :
        self.setWindowTitle("未授权访问扫描工具")
        self.resize(1200, 700)

        # 主布局：左侧服务设置 + 右侧扫描区域
        main_splitter = QSplitter(Qt.Orientation.Horizontal)

        # 左侧：服务与端口设置
        left_widget = QWidget()
        left_layout = QVBoxLayout(left_widget)
        left_layout.setContentsMargins(5, 5, 5, 5)
        left_layout.setSpacing(5)

        # 服务表格
        service_group = QGroupBox("服务与端口设置（可编辑）")
        service_layout = QVBoxLayout(service_group)
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        table_container = QWidget()
        table_layout = QVBoxLayout(table_container)

        self.service_table = QTableWidget()
        self.service_table.setColumnCount(2)
        self.service_table.setHorizontalHeaderLabels(["服务", "端口"])
        self.service_table.horizontalHeader().setStretchLastSection(True)
        self.service_table.setMinimumWidth(300)

        # 填充服务数据
        services = sorted(VulnerabilityScanner().default_ports.items(), key=lambda x : x[0])
        self.service_table.setRowCount(len(services))
        for row, (name, port) in enumerate(services) :
            item_name = QTableWidgetItem(name)
            item_name.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsUserCheckable)
            item_name.setCheckState(Qt.CheckState.Checked)
            self.service_table.setItem(row, 0, item_name)

            item_port = QTableWidgetItem(str(port))
            item_port.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsEditable)
            self.service_table.setItem(row, 1, item_port)

        table_layout.addWidget(self.service_table)
        scroll_area.setWidget(table_container)
        service_layout.addWidget(scroll_area)

        # 服务表格控制按钮
        btn_layout = QHBoxLayout()
        self.select_all_btn = QPushButton("全选")
        self.deselect_all_btn = QPushButton("取消全选")
        self.select_all_btn.clicked.connect(lambda : self.set_all_check_state(Qt.CheckState.Checked))
        self.deselect_all_btn.clicked.connect(lambda : self.set_all_check_state(Qt.CheckState.Unchecked))
        btn_layout.addWidget(self.select_all_btn)
        btn_layout.addWidget(self.deselect_all_btn)
        service_layout.addLayout(btn_layout)
        left_layout.addWidget(service_group, 1)
        main_splitter.addWidget(left_widget)

        # 右侧：扫描控制与结果展示
        right_widget = QWidget()
        right_layout = QVBoxLayout(right_widget)

        # 目标输入区域
        target_group = QGroupBox("扫描目标")
        target_layout = QVBoxLayout()
        target_input_layout = QHBoxLayout()
        self.target_edit = QLineEdit()
        self.target_edit.setPlaceholderText("支持格式：IP、IP:PORT、URL、域名（多个用逗号分隔）")
        self.load_file_btn = QPushButton("从文件导入")
        self.load_file_btn.clicked.connect(self.load_targets_from_file)
        target_input_layout.addWidget(self.target_edit)
        target_input_layout.addWidget(self.load_file_btn)
        target_layout.addLayout(target_input_layout)
        target_group.setLayout(target_layout)
        right_layout.addWidget(target_group)

        # 扫描设置
        settings_group = QGroupBox("扫描设置")
        settings_layout = QHBoxLayout()
        self.thread_spin = QSpinBox()
        self.thread_spin.setRange(1, 100)  # 提高线程上限
        self.thread_spin.setValue(20)
        settings_layout.addWidget(QLabel("线程数:"))
        settings_layout.addWidget(self.thread_spin)

        self.timeout_spin = QSpinBox()
        self.timeout_spin.setRange(1, 10)
        self.timeout_spin.setValue(5)
        self.timeout_spin.setSuffix("秒")
        settings_layout.addWidget(QLabel("超时时间:"))
        settings_layout.addWidget(self.timeout_spin)

        settings_layout.addSpacing(20)
        self.proxy_edit = QLineEdit()
        self.proxy_edit.setPlaceholderText("代理（如http(socks5)://127.0.0.1:1080）")
        settings_layout.addWidget(QLabel("代理:"))
        settings_layout.addWidget(self.proxy_edit)
        settings_group.setLayout(settings_layout)
        right_layout.addWidget(settings_group)

        # 控制按钮
        control_layout = QHBoxLayout()
        self.start_btn = QPushButton("开始扫描")
        self.stop_btn = QPushButton("停止扫描")
        self.export_btn = QPushButton("导出结果")
        self.clear_btn = QPushButton("清空结果")
        self.start_btn.clicked.connect(self.start_scan)
        self.stop_btn.clicked.connect(self.stop_scan)
        self.export_btn.clicked.connect(self.export_results)
        self.clear_btn.clicked.connect(self.clear_results)
        self.stop_btn.setEnabled(False)
        control_layout.addWidget(self.start_btn)
        control_layout.addWidget(self.stop_btn)
        control_layout.addWidget(self.export_btn)
        control_layout.addWidget(self.clear_btn)
        right_layout.addLayout(control_layout)

        # 进度条
        self.progress_bar = QProgressBar()
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setFormat("扫描进度: %p%")
        right_layout.addWidget(self.progress_bar)

        # 结果显示
        result_group = QGroupBox("扫描日志与结果")
        result_layout = QVBoxLayout()
        self.result_display = QTextEdit()
        self.result_display.setReadOnly(True)
        result_layout.addWidget(self.result_display)
        result_group.setLayout(result_layout)
        right_layout.addWidget(result_group, 1)

        main_splitter.addWidget(right_widget)
        main_splitter.setSizes([350, 850])

        # 主窗口布局
        main_layout = QVBoxLayout(self)
        main_layout.addWidget(main_splitter)

        # 保存结果数据
        self.scan_results = []

    # ---------------------- UI功能方法 ----------------------
    def set_all_check_state(self, state) :
        for row in range(self.service_table.rowCount()) :
            item = self.service_table.item(row, 0)
            item.setCheckState(state)

    def get_selected_services(self) :
        services = []
        custom_ports = {}
        for row in range(self.service_table.rowCount()) :
            name_item = self.service_table.item(row, 0)
            if name_item.checkState() == Qt.CheckState.Checked :
                service_name = name_item.text()
                services.append(service_name)
                # 读取用户自定义端口
                try :
                    port = int(self.service_table.item(row, 1).text())
                    custom_ports[service_name] = port
                except ValueError :
                    pass  # 无效端口则使用默认
        return services, custom_ports

    def load_targets_from_file(self) :
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择目标文件", "", "文本文件 (*.txt);;所有文件 (*)"
        )
        if file_path :
            try :
                with open(file_path, 'r') as f :
                    targets = [line.strip() for line in f if line.strip()]
                self.target_edit.setText(",".join(targets))
            except Exception as e :
                QMessageBox.warning(self, "错误", f"加载文件失败: {str(e)}")

    def start_scan(self) :
        # 1. 解析目标
        target_text = self.target_edit.text().strip()
        if not target_text :
            QMessageBox.warning(self, "警告", "请输入目标")
            return
        targets_raw = [t.strip() for t in target_text.split(',') if t.strip()]
        # 标准化所有目标（过滤无效目标）
        targets = []
        for t in targets_raw :
            normalized = normalize_target_input(t)
            if normalized :
                targets.append(normalized)
            else :
                self.result_display.append(f"[{datetime.now().strftime('%H:%M:%S')}] 无效目标: {t}")
        if not targets :
            QMessageBox.warning(self, "警告", "无有效目标（请检查输入格式）")
            return

        # 2. 获取选中的服务
        services, custom_ports = self.get_selected_services()
        if not services :
            QMessageBox.warning(self, "警告", "请至少选择一个服务")
            return

        # 3. 初始化扫描
        self.start_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self.result_display.clear()
        self.scan_results = []
        self.progress_bar.setValue(0)
        self.result_display.append(
            f"[{datetime.now().strftime('%H:%M:%S')}] 开始扫描，目标数: {len(targets)}，服务数: {len(services)}")

        # 4. 启动扫描线程
        proxy = self.proxy_edit.text().strip() or None
        self.scan_thread = ScanThread(
            targets=targets,
            services=services,
            custom_ports=custom_ports,
            proxy=proxy,
            threads=self.thread_spin.value()
        )
        # 绑定信号与槽函数
        self.scan_thread.scanning_service.connect(self.log_scanning)
        self.scan_thread.service_result.connect(self.log_service_result)
        self.scan_thread.progress_updated.connect(self.progress_bar.setValue)
        self.scan_thread.scan_completed.connect(self.on_scan_completed)
        self.scan_thread.start()

    def stop_scan(self) :
        if self.scan_thread and self.scan_thread.isRunning() :
            self.scan_thread.stop()
            self.stop_btn.setEnabled(False)
            self.result_display.append(f"[{datetime.now().strftime('%H:%M:%S')}] 已强制停止所有扫描任务")

    def log_scanning(self, target, service, path) :
        """记录正在扫描的服务日志"""
        self.result_display.append(
            f"[{datetime.now().strftime('%H:%M:%S')}] 正在扫描: {target} - {service}（路径: {path}）")
        self.result_display.moveCursor(QTextCursor.MoveOperation.End)

    def log_service_result(self, target, service, status, message) :
        """记录服务扫描结果"""
        prefix = "✅ 发现漏洞" if status else "❌ 未发现漏洞"
        self.result_display.append(
            f"[{datetime.now().strftime('%H:%M:%S')}] {prefix}: {target} - {service} - {message}")
        self.result_display.moveCursor(QTextCursor.MoveOperation.End)

    def on_scan_completed(self, results) :
        """扫描完成后汇总结果，使用服务对应端口显示目标URL"""
        self.scan_results = results
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.progress_bar.setValue(100)

        # 获取选中服务的自定义端口和scanner实例
        _, custom_ports = self.get_selected_services()
        scanner = VulnerabilityScanner()

        # 汇总漏洞数量
        total_vulns = 0
        vuln_details = []
        for result in results :
            # 解析原始目标信息
            target_info = normalize_target_input(result["target"])
            if not target_info :
                continue

            vulns = [v for v in result["vulnerabilities"] if v["status"]]
            total_vulns += len(vulns)
            if vulns :
                # 按服务分组显示（同一目标不同服务可能对应不同端口）
                service_groups = {}
                for v in vulns :
                    # 使用现有方法获取服务实际URL
                    service_url = scanner.get_display_url(target_info, v["service"], custom_ports)
                    if service_url not in service_groups :
                        service_groups[service_url] = []
                    service_groups[service_url].append(v)

                # 按服务URL分组添加到汇总信息
                for service_url, group_vulns in service_groups.items() :
                    vuln_details.append(f"目标: {service_url}")
                    for v in group_vulns :
                        vuln_details.append(f"  - 服务: {v['service']}，信息: {v['message']}")

        # 显示汇总信息
        self.result_display.append("\n" + "=" * 50)
        self.result_display.append(f"[{datetime.now().strftime('%H:%M:%S')}] 扫描完成！总漏洞数: {total_vulns}")
        if vuln_details :
            self.result_display.append("存在未授权访问的服务列表：")
            self.result_display.append("\n".join(vuln_details))
        else :
            self.result_display.append("未发现任何未授权访问漏洞")
        self.result_display.append("=" * 50)
        self.result_display.moveCursor(QTextCursor.MoveOperation.End)

    def export_results(self) :
        if not self.scan_results :
            QMessageBox.warning(self, "警告", "没有可导出的结果")
            return

        # 添加Excel选项到文件对话框
        file_path, _ = QFileDialog.getSaveFileName(
            self, "保存结果", f"unauth_scan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel文件 (*.xlsx);;JSON文件 (*.json);;文本文件 (*.txt);;所有文件 (*)"
        )

        if file_path :
            try :
                # 获取选中服务的自定义端口配置
                _, custom_ports = self.get_selected_services()
                # 创建scanner实例用于调用get_display_url
                scanner = VulnerabilityScanner()

                # 导出为Excel文件
                if file_path.endswith('.xlsx') :
                    self._export_to_excel(file_path, scanner, custom_ports)
                # 导出为JSON文件
                elif file_path.endswith('.json') :
                    with open(file_path, 'w', encoding='utf-8') as f :
                        json.dump(self.scan_results, f, indent=2, ensure_ascii=False)
                # 导出为文本文件
                else :
                    with open(file_path, 'w', encoding='utf-8') as f :
                        f.write(f"未授权访问扫描报告 - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                        total_vulns = sum(
                            len([v for v in r['vulnerabilities'] if v['status']]) for r in self.scan_results)
                        f.write(f"目标数量: {len(self.scan_results)}, 总漏洞数: {total_vulns}\n\n")

                        for result in self.scan_results :
                            # 解析原始目标信息
                            target_info = normalize_target_input(result['target'])
                            f.write(f"目标: {result['target']}\n")

                            for vuln in result["vulnerabilities"] :
                                # 使用现有方法获取服务实际URL
                                service_url = scanner.get_display_url(target_info, vuln['service'], custom_ports)
                                status = "存在漏洞" if vuln["status"] else "安全"
                                f.write(f"  - {vuln['service']} ({service_url}): {status} - {vuln['message']}\n")
                            f.write("\n")

                QMessageBox.information(self, "成功", f"结果已导出到 {file_path}")
            except Exception as e :
                QMessageBox.warning(self, "错误", f"导出失败: {str(e)}")

    def _export_to_excel(self, file_path, scanner, custom_ports) :
        """将扫描结果导出为Excel格式，使用现有方法获取正确URL"""
        # 创建工作簿和工作表
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "扫描结果"

        # 定义样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        highlight_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 存在漏洞的行高亮

        # 设置表头
        headers = ["序号", "目标URL", "服务名称", "是否存在漏洞", "漏洞信息", "扫描时间"]
        for col, header in enumerate(headers, 1) :
            cell = worksheet.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 填充数据
        row_num = 2
        index = 1
        scan_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        for result in self.scan_results :
            # 解析原始目标信息用于生成正确URL
            target_info = normalize_target_input(result["target"])
            for vuln in result["vulnerabilities"] :
                # 使用现有方法获取服务实际扫描的URL（包含正确端口）
                service_url = scanner.get_display_url(target_info, vuln['service'], custom_ports)

                # 序号
                cell = worksheet.cell(row=row_num, column=1)
                cell.value = index
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")

                # 目标URL - 使用服务实际端口的URL
                cell = worksheet.cell(row=row_num, column=2)
                cell.value = service_url
                cell.border = border
                cell.alignment = Alignment(vertical="center")

                # 服务名称
                cell = worksheet.cell(row=row_num, column=3)
                cell.value = vuln["service"]
                cell.border = border
                cell.alignment = Alignment(vertical="center")

                # 是否存在漏洞
                cell = worksheet.cell(row=row_num, column=4)
                status_text = "是" if vuln["status"] else "否"
                cell.value = status_text
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")

                # 如果存在漏洞，整行高亮
                if vuln["status"] :
                    for col in range(1, len(headers) + 1) :
                        worksheet.cell(row=row_num, column=col).fill = highlight_fill

                # 漏洞信息
                cell = worksheet.cell(row=row_num, column=5)
                cell.value = vuln["message"]
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical="center")

                # 扫描时间
                cell = worksheet.cell(row=row_num, column=6)
                cell.value = scan_time
                cell.border = border
                cell.alignment = Alignment(vertical="center")

                row_num += 1
                index += 1

        # 调整列宽
        column_widths = [8, 35, 15, 12, 50, 20]  # 增加URL列宽以显示完整地址
        for col in range(1, len(headers) + 1) :
            worksheet.column_dimensions[get_column_letter(col)].width = column_widths[col - 1]

        # 添加自动筛选
        worksheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_num - 1}"

        # 保存文件
        workbook.save(file_path)

    def clear_results(self) :
        self.result_display.clear()
        self.scan_results = []
        self.progress_bar.setValue(0)


if __name__ == "__main__":
    signal.signal(signal.SIGINT, signal.SIG_DFL)
    app = QApplication(sys.argv)
    window = UnauthorizedScanTab()
    window.show()
    sys.exit(app.exec())
